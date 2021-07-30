"""概要(Summary line)

作業者が必要なバーコードのリストをシートごとに作成する

Example
-------
python create_xlsx.py [f:filename]

"""

import os

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side


SHEET_A4_WIDTH = 100    # A4での幅の最大限,プリンターによって若干異なる.実機は100でPDFは105
SHEET_A4_NROW = 21      # A4での行数,プリンターによって若干異なる.実機は19でPDFは20
# PDFだと見切れてしまう、実機の設定が安牌っぽい

USE_LOCATION_IDS = ('2A-1B-01', '2A-1B-09', '2A-1B-14', '2A-1C-01', '2A-1C-11', '2A-1E-02', '2A-1E-13', '2A-1F-03', '2A-1F-06', '2A-1F-07', '2A-1F-10', '2A-1F-12', '2A-1G-04', '2A-1G-06', '2A-1G-08', '2A-1H-01', '2A-1H-07', '2A-1I-04', '2A-1I-06', '2A-1I-08', '2A-1J-04', '2A-1K-13', '2A-1L-02', '2A-1L-08', '2A-1L-14', '2A-1M-11', '2A-2A-13', '2A-2A-14', '2A-2A-15', '2A-2A-19', '2A-2B-02', '2A-2B-05', '2A-2B-09', '2A-2B-13', '2A-2B-18', '2A-2C-06', '2A-2C-11', '2A-2C-15', '2A-2D-01', '2A-2D-05', '2A-2D-07', '2A-2D-09', '2A-2D-10', '2A-2D-15', '2A-2D-17', '2A-2E-05', '2A-2E-08', '2A-2E-12', '2A-2F-03', '2A-2F-06', '2A-2F-09', '2A-2G-01', '2A-2H-00', '2A-2H-12', '2A-2H-14', '2A-2I-01', '2A-2I-17', '2A-2J-01', '2A-2J-08', '2A-2J-15', '2A-2J-17', '2A-2K-05', '2A-2K-10', '2A-2K-15', '2A-2L-05', '2A-2L-07', '2A-2L-08', '2A-2L-09', '2A-3A-05', '2A-3B-06', '2A-3B-11', '2A-3B-12', '2A-3C-15', '2A-3D-05', '2A-3D-08', '2A-3D-13', '2A-3D-16', '2A-3E-15', '2A-3F-04', '2A-3F-12', '2A-3F-13', '2A-3F-14', '2A-3G-01', '2A-3G-06', '2A-3G-08', '2A-3G-10', '2A-3G-11', '2A-3H-01', '2A-3H-06', '2A-3H-09', '2A-3H-11', '2A-3I-03', '2A-3I-09', '2A-3J-04', '2A-3J-09', '2A-3J-11', '2A-3K-02', '2A-3K-09', '2A-3K-13', '2A-3K-14', '2A-3L-01', '2A-3L-08', '2A-3L-11', '2A-3L-13', '2A-3N-02', '2A-3N-04', '2A-3N-10', '2A-3N-13', '2A-3N-16', '2A-3O-03', '2A-3O-06', '2B-1A-01', '2B-1A-03', '2B-1B-03', '2B-1C-01', '2B-1C-02', '2B-1C-12', '2B-1D-03', '2B-1E-06', '2B-1E-09', '2B-1G-07', '2B-1H-11', '2B-2A-01', '2B-2A-06', '2B-2B-09', '2B-2B-16', '2B-2C-12', '2B-2C-15', '2B-2D-03', '2B-2E-00', '2B-2F-01', '2B-2F-06', '2B-2F-08', '2B-2H-03', '2B-2H-16', '2B-2I-12', '2B-2I-13', '2B-2J-04', '2B-2J-07', '2B-2J-09', '2B-2K-02', '2B-2L-03', '2B-3A-04', '2B-3A-14', '2B-3B-01', '2B-3B-09', '2B-3C-01', '2B-3C-07', '2B-3D-09', '2B-3D-16', '2B-3E-03', '2B-3E-06', '2B-3F-04', '2B-3F-05', '2B-3F-13', '2B-3G-09', '2B-3I-05', '2B-3J-03', '2B-3K-03', '2B-3K-07', '2B-3K-10', '2B-3L-06', '2B-3L-13', '2B-3M-15', '2B-3N-02', '2B-3N-11', '2B-3N-12', '2B-3O-02', '2C-1A-04', '2C-1A-08', '2C-1A-09', '2C-1A-10', '2C-1B-04', '2C-1B-12', '2C-1B-13', '2C-1B-14', '2C-1C-02', '2C-1C-09', '2C-1C-15', '2C-1D-04', '2C-1D-05', '2C-1D-06', '2C-1E-04', '2C-1E-06', '2C-1F-05', '2C-1F-07', '2C-1F-11', '2C-1G-03', '2C-1H-09', '2C-1H-10', '2C-1H-13', '2C-1I-08', '2C-1J-03', '2C-1J-04', '2C-1J-06', '2C-1J-07', '2C-1K-03', '2C-1L-08', '2C-2A-16', '2C-2B-02', '2C-2B-03', '2C-2B-13', '2C-2B-15', '2C-2B-17', '2C-2E-09', '2C-2H-17', '2C-2I-10', '2C-2J-13', '2C-2J-15', '2C-2K-09', '2C-2K-14', '2C-2L-03', '2C-2L-06', '2C-2L-08', '2C-3C-03', '2C-3D-06', '2C-3D-11', '2C-3D-15', '2C-3E-07', '2C-3H-14', '2C-3I-06', '2C-3I-10', '2C-3I-12', '2C-3I-15', '2C-3J-05', '2D-1A-01', '2D-1A-02', '2D-1A-08', '2D-1A-10', '2D-1A-11', '2D-1A-13', '2D-1B-01', '2D-1B-11', '2D-1C-02', '2D-1C-06', '2D-1C-07', '2D-1C-09', '2D-1C-10', '2D-1C-15')
# '2A-2H-00'と'2B-2E-00'がlocation_list.csvにない


def create_xlsx_file(barcodefile="pos.csv", wmsfile="order_list.txt", dstfile="wms.xlsx"):
    """概要(Summary line)

    印刷用エクセルファイルを作成する

    Parameters
    ----------
    barcodefile : str
        バーコードとラベルが記されたファイル
    wmsfile : str
        作業者と作業内容が記されたファイル
    dstfile : str
        出力ファイル名

    Returns
    -------
    int
        ファイル作成の成功(0)or失敗

    """

    # barcodeファイルの読み込み
    barcode_dic = {}    # キーはラベル、値はバーコード
    with open(barcodefile, "r") as fr1:
        line = fr1.readline()
        while line:
            if line[-1] == "\n":
                line = line[:-1]
            line_split = line.split(",")
            barcode = "*" + line_split[-2] + "*"
            label = line_split[-1]
            barcode_dic[label] = barcode
            line = fr1.readline()
    # wms_file読み込み
    wms_dic = {}    # キーはworkerID、値は作業リスト
    with open(wmsfile, "r") as fr2:
        line = fr2.readline()
        while line:
            if line[-1] == "\n":
                line = line[:-1]
            line_split = line.split(":")
            worker_id = line_split[0]
            work_list = line_split[1]
            wms_dic[worker_id] = work_list.split(",")
            line = fr2.readline()
    # ワークブック作成
    wb = openpyxl.Workbook()
    for worker_id in wms_dic:
        wb.create_sheet("worker" + str(worker_id))
        ws = wb.worksheets[-1]
        __create_sheet(ws, worker_id, barcode_dic, wms_dic[worker_id])
    # エクセルファイル保存
    wb.remove(wb.worksheets[0])
    wb.active = wb.worksheets[0]
    if os.path.splitext(dstfile)[1] != ".xlsx":
        dstfile += ".xlsx"
    try:
        wb.save(dstfile)
    except(PermissionError):
        return 3

    return 0


def __create_sheet(ws, worker_id, barcode_dic, work_list):
    """概要(Summary line)

    作業者ごとのリストを作成する

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        対象のワークシート
    worker_id : str
        ワーカーID
    barcode_dic : dict
        キーはラベル、値はバーコードの辞書
    work_list : list
        workerIDが行うラベルのリスト

    Returns
    -------
    None

    """

    # ヘッダー書き込み
    title_cell = ws.cell(2, 2)
    title_cell.value = "作業者" + str(worker_id) + "用リスト"
    title_cell.font = Font(size=20)
    number_font = Font(size=18)
    label_font = Font(size=20)
    barcode_font = Font(name="CODE39", size=18)
    for i, worker_label in enumerate(work_list):
        row_idx = 2*i + 4
        cell_number = ws.cell(row_idx, 2)
        cell_label = ws.cell(row_idx, 4)
        if i % 2 == 0:
            cell_barcode = ws.cell(row_idx, 6)
        else:
            cell_barcode = ws.cell(row_idx, 7)
        cell_number.value = str(i + 1)
        cell_label.value = worker_label
        if worker_label in barcode_dic:
            cell_barcode.value = barcode_dic[worker_label]
        cell_barcode.font = barcode_font
        ws.row_dimensions[row_idx].height = 24
        ws.row_dimensions[row_idx + 1].height = 22
    # セルの書式設定
    # ws.column_dimensions["A"].width = 10
    # ws.column_dimensions["D"].width = 5
    # ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20