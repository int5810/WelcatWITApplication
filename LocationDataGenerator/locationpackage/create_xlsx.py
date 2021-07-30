"""概要(Summary line)

バーコードを付与したワークブックを作成する

Example
-------
python create_xlsx.py [f:filename]

"""

import os

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side


SHEET_A4_WIDTH = 100    # A4での幅の最大限,プリンターによって若干異なる.実機は100でPDFは105
SHEET_A4_NROW = 21      # A4での行数,プリンターによって若干異なる.実機は19でPDFは20
# PDFだと見切れてしまう、実機の設定が安牌っぽい

USE_LOCATION_IDS = ('2A-1B-01', '2A-1B-09', '2A-1B-14', '2A-1C-01', '2A-1C-11', '2A-1E-02', '2A-1E-13', '2A-1F-03', '2A-1F-06', '2A-1F-07', '2A-1F-10', '2A-1F-12', '2A-1G-04', '2A-1G-06', '2A-1G-08', '2A-1H-01', '2A-1H-07', '2A-1I-04', '2A-1I-06', '2A-1I-08', '2A-1J-04', '2A-1K-13', '2A-1L-02', '2A-1L-08', '2A-1L-14', '2A-1M-11', '2A-2A-13', '2A-2A-14', '2A-2A-15', '2A-2A-19', '2A-2B-02', '2A-2B-05', '2A-2B-09', '2A-2B-13', '2A-2B-18', '2A-2C-06', '2A-2C-11', '2A-2C-15', '2A-2D-01', '2A-2D-05', '2A-2D-07', '2A-2D-09', '2A-2D-10', '2A-2D-15', '2A-2D-17', '2A-2E-05', '2A-2E-08', '2A-2E-12', '2A-2F-03', '2A-2F-06', '2A-2F-09', '2A-2G-01', '2A-2H-00', '2A-2H-12', '2A-2H-14', '2A-2I-01', '2A-2I-17', '2A-2J-01', '2A-2J-08', '2A-2J-15', '2A-2J-17', '2A-2K-05', '2A-2K-10', '2A-2K-15', '2A-2L-05', '2A-2L-07', '2A-2L-08', '2A-2L-09', '2A-3A-05', '2A-3B-06', '2A-3B-11', '2A-3B-12', '2A-3C-15', '2A-3D-05', '2A-3D-08', '2A-3D-13', '2A-3D-16', '2A-3E-15', '2A-3F-04', '2A-3F-12', '2A-3F-13', '2A-3F-14', '2A-3G-01', '2A-3G-06', '2A-3G-08', '2A-3G-10', '2A-3G-11', '2A-3H-01', '2A-3H-06', '2A-3H-09', '2A-3H-11', '2A-3I-03', '2A-3I-09', '2A-3J-04', '2A-3J-09', '2A-3J-11', '2A-3K-02', '2A-3K-09', '2A-3K-13', '2A-3K-14', '2A-3L-01', '2A-3L-08', '2A-3L-11', '2A-3L-13', '2A-3N-02', '2A-3N-04', '2A-3N-10', '2A-3N-13', '2A-3N-16', '2A-3O-03', '2A-3O-06', '2B-1A-01', '2B-1A-03', '2B-1B-03', '2B-1C-01', '2B-1C-02', '2B-1C-12', '2B-1D-03', '2B-1E-06', '2B-1E-09', '2B-1G-07', '2B-1H-11', '2B-2A-01', '2B-2A-06', '2B-2B-09', '2B-2B-16', '2B-2C-12', '2B-2C-15', '2B-2D-03', '2B-2E-00', '2B-2F-01', '2B-2F-06', '2B-2F-08', '2B-2H-03', '2B-2H-16', '2B-2I-12', '2B-2I-13', '2B-2J-04', '2B-2J-07', '2B-2J-09', '2B-2K-02', '2B-2L-03', '2B-3A-04', '2B-3A-14', '2B-3B-01', '2B-3B-09', '2B-3C-01', '2B-3C-07', '2B-3D-09', '2B-3D-16', '2B-3E-03', '2B-3E-06', '2B-3F-04', '2B-3F-05', '2B-3F-13', '2B-3G-09', '2B-3I-05', '2B-3J-03', '2B-3K-03', '2B-3K-07', '2B-3K-10', '2B-3L-06', '2B-3L-13', '2B-3M-15', '2B-3N-02', '2B-3N-11', '2B-3N-12', '2B-3O-02', '2C-1A-04', '2C-1A-08', '2C-1A-09', '2C-1A-10', '2C-1B-04', '2C-1B-12', '2C-1B-13', '2C-1B-14', '2C-1C-02', '2C-1C-09', '2C-1C-15', '2C-1D-04', '2C-1D-05', '2C-1D-06', '2C-1E-04', '2C-1E-06', '2C-1F-05', '2C-1F-07', '2C-1F-11', '2C-1G-03', '2C-1H-09', '2C-1H-10', '2C-1H-13', '2C-1I-08', '2C-1J-03', '2C-1J-04', '2C-1J-06', '2C-1J-07', '2C-1K-03', '2C-1L-08', '2C-2A-16', '2C-2B-02', '2C-2B-03', '2C-2B-13', '2C-2B-15', '2C-2B-17', '2C-2E-09', '2C-2H-17', '2C-2I-10', '2C-2J-13', '2C-2J-15', '2C-2K-09', '2C-2K-14', '2C-2L-03', '2C-2L-06', '2C-2L-08', '2C-3C-03', '2C-3D-06', '2C-3D-11', '2C-3D-15', '2C-3E-07', '2C-3H-14', '2C-3I-06', '2C-3I-10', '2C-3I-12', '2C-3I-15', '2C-3J-05', '2D-1A-01', '2D-1A-02', '2D-1A-08', '2D-1A-10', '2D-1A-11', '2D-1A-13', '2D-1B-01', '2D-1B-11', '2D-1C-02', '2D-1C-06', '2D-1C-07', '2D-1C-09', '2D-1C-10', '2D-1C-15')
# '2A-2H-00'と'2B-2E-00'がlocation_list.csvにない


def create_xlsx_file(srcfile="pos.csv", dstfile="pos.xlsx"):
    """概要(Summary line)

    印刷用エクセルファイルを作成する

    Parameters
    ----------
    filename : str
        入力csvファイル名
    dstfile : str
        出力ファイル名

    Returns
    -------
    int
        ファイル作成の成功(0)or失敗

    """

    # ファイルの読み込み
    csv_data = []
    locations = []
    with open(srcfile, "r") as fr:
        line = fr.readline()
        while line:
            line_split = line[:-1].split(",")
            if True:#line_split[4] in USE_LOCATION_IDS:   # True
                locations.append(line_split[4])
                csv_data.append([line_split[0], line_split[1], line_split[2], line_split[3], line_split[4]])
            line = fr.readline()
    # ワークブックと各種シートを作成
    wb = openpyxl.Workbook()
    if not __create_summary_sheet(wb, csv_data):
        return 1
    if not __create_print_sheets(wb, csv_data):
        return 2
    # 保存
    wb.active = wb.worksheets[0]
    if os.path.splitext(dstfile)[1] != ".xlsx":
        dstfile += ".xlsx"
    try:
        wb.save(dstfile)
    except(PermissionError):
        return 3

    return 0


def __create_summary_sheet(wb, csv_data):
    """概要(Summary line)

    要約リストのシートを作成する

    Parameters
    ----------
    wb : openpyxl.Workbook
        対象のワークブック
    csv_data : list
        csvファイルのデータ

    Returns
    -------
    bool
        シート作成の成功or失敗

    """

    # ヘッダー書き込み
    ws = wb.worksheets[0]
    ws.title = "バーコード全リスト"
    row_idx = 1 # 1オリジンのインデクサ
    cell_b = ws.cell(row_idx, 2)
    cell_c = ws.cell(row_idx, 3)
    cell_d = ws.cell(row_idx, 4)
    cell_e = ws.cell(row_idx, 5)
    cell_f = ws.cell(row_idx, 6)
    cell_b.value = "X"
    cell_c.value = "Y"
    cell_d.value = "Z"
    cell_e.value = "バーコードの値"
    cell_f.value = "バーコード"
    cell_b.alignment = Alignment(horizontal="right")
    cell_c.alignment = Alignment(horizontal="right")
    cell_d.alignment = Alignment(horizontal="right")
    cell_e.alignment = Alignment(horizontal="center")
    # バーコードリストデータ入力
    normal_alignment = Alignment(vertical="center")
    normal_font = Font(size=12)
    for csv_row in csv_data:
        row_idx += 1
        cell_a = ws.cell(row_idx, 1)
        cell_b = ws.cell(row_idx, 2)
        cell_c = ws.cell(row_idx, 3)
        cell_d = ws.cell(row_idx, 4)
        cell_e = ws.cell(row_idx, 5)
        cell_f = ws.cell(row_idx, 6)
        cell_a.value = csv_row[4]
        cell_b.value = float(csv_row[0])
        cell_c.value = float(csv_row[1])
        cell_d.value = float(csv_row[2])
        cell_e.value = csv_row[3]
        # '=CONCATENATE("(実験中) ",バーコード全リスト!A{0}," : ",バーコード全リスト!E{0})'.format(list_row_idx)
        # cell_f.value = "=E" + str(row_idx)
        cell_f.value = '=CONCATENATE("*",E' + str(row_idx) + ',"*")'
        cell_a.alignment = normal_alignment
        cell_b.alignment = normal_alignment
        cell_c.alignment = normal_alignment
        cell_d.alignment = normal_alignment
        cell_e.alignment = Alignment(horizontal="center", vertical="center")
        cell_f.alignment = normal_alignment
        cell_a.font = normal_font
        cell_b.font = normal_font
        cell_c.font = normal_font
        cell_d.font = normal_font
        cell_e.font = normal_font
        cell_f.font = Font(name="CODE39", size=28)
        ws.row_dimensions[row_idx].height = 22
    # セルの書式設定
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 31
    # ws.column_dimensions["E"].alignment = Alignment(horizontal="center")#できないっぽい
    # ws.column_dimensions["F"].font = Font(name="CODE39")#できないっぽい

    return True


def __create_print_sheets(wb, csv_data, col_num=3):
    """概要(Summary line)

    各バーコード専用のシートを作成する

    Parameters
    ----------
    wb : openpyxl.Workbook
        対象のワークブック
    csv_data : list
        csvファイルのデータ
    col_num : int
        バーコードの列数、最大26

    Returns
    -------
    bool
        シート作成の成功or失敗

    """

    start_idx = 0
    while len(csv_data) != 0:
        wb.create_sheet("印刷用" + str(len(csv_data)))   # シート名は一時的に付ける
        ws = wb.worksheets[-1]
        add_len = min(len(csv_data), SHEET_A4_NROW * col_num)
        start_label = csv_data[0][4]
        end_label = csv_data[add_len - 1][4]
        __create_print_sheet(ws, csv_data[:add_len], col_num, start_idx)
        csv_data = csv_data[add_len:]
        ws.title = "{}～{}".format(start_label, end_label)
        start_idx += add_len

    return True


def __create_print_sheet(ws, csv_data, col_num, start_idx):
    """概要(Summary line)

    各バーコード専用のシートを作成する

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        対象のワークシート
    csv_data : list
        csvファイルのデータ
    col_num : int
        バーコードの列数、最大26

    Returns
    -------
    bool
        シート作成の成功or失敗

    """

    # csvデータを入力
    row_idx = 1
    col_idx = 1
    top_alignment = Alignment(horizontal="center", vertical="top")
    bottom_left_alignment = Alignment(horizontal="right", vertical="bottom")
    bottom_right_alignment = Alignment(horizontal="center", vertical="bottom")
    top_font = Font(name="CODE39", size=__define_top_fontsize(col_num))
    bottom_left_font = Font(size=__define_bottom_fontsize(col_num), color="FF0000", italic=True)
    bottom_right_font = Font(size=__define_bottom_fontsize(col_num))
    top_height = 24
    bottom_height = 13
    top_margin_height = 3
    middle_margin_height = 0.5
    bottom_margin_height = 0.5
    side = Side(style="thin")
    top_margin_left_border = Border(top=side, left=side)
    top_margin_right_border = Border(top=side, right=side)
    middle_left_border = Border(left=side)
    middle_right_border = Border(right=side)
    bottom_border = Border(bottom=side, left=side, right=side)
    bottom_left_border = Border(bottom=side, left=side)
    bottom_right_border = Border(bottom=side, right=side)
    for i in range(len(csv_data)):
        list_row_idx = i + 2
        # セル定義
        cell_top_margin_left = ws.cell(row_idx, 2*col_idx - 1)
        cell_top_margin_right = ws.cell(row_idx, 2 * col_idx)
        cell_top_left = ws.cell(row_idx+1, 2*col_idx - 1)
        cell_top_right = ws.cell(row_idx+1, 2 * col_idx)
        cell_top_2right = ws.cell(row_idx+1, 2*col_idx + 1)  # 結合セルの右側罫線がいじれないので、もう1つ右のセルをいじる
        cell_middle_margin_left = ws.cell(row_idx+2, 2*col_idx - 1)
        cell_middle_margin_right = ws.cell(row_idx+2, 2 * col_idx)
        cell_bottom_left = ws.cell(row_idx+3, 2*col_idx - 1)
        cell_bottom_right = ws.cell(row_idx+3, 2 * col_idx)
        cell_bottom_margin_left = ws.cell(row_idx+4, 2*col_idx - 1)
        cell_bottom_margin_right = ws.cell(row_idx+4, 2 * col_idx)
        # 値入力
        cell_top_left.value = '=CONCATENATE("*",バーコード全リスト!E' + str(list_row_idx + start_idx) + ',"*")'
        cell_bottom_left.value = "実験中"
        cell_bottom_right.value = '=CONCATENATE(バーコード全リスト!A{0}," : ",バーコード全リスト!E{0})'.format(list_row_idx + start_idx)
        # アライメント
        cell_top_left.alignment = top_alignment
        cell_bottom_left.alignment = bottom_left_alignment
        cell_bottom_right.alignment = bottom_right_alignment
        # フォント
        cell_top_left.font = top_font
        cell_bottom_left.font = bottom_left_font
        cell_bottom_right.font = bottom_right_font
        # 罫線
        cell_top_margin_left.border = top_margin_left_border
        cell_top_margin_right.border = top_margin_right_border
        cell_top_left.border = middle_left_border
        cell_top_right.border = middle_right_border
        cell_middle_margin_left.border = middle_left_border
        cell_middle_margin_right.border = middle_right_border
        cell_bottom_left.border = middle_left_border
        cell_bottom_right.border = middle_right_border
        cell_bottom_margin_left.border = bottom_left_border
        cell_bottom_margin_right.border = bottom_right_border
        # セルの結合
        left_char = __convert_num2xlscol(2*col_idx - 2)
        right_char = __convert_num2xlscol(2*col_idx - 1)
        ws.merge_cells(left_char + str(row_idx + 1) + ":" + right_char + str(row_idx + 1))
        # 高さ調整
        if col_idx == 1:
            ws.row_dimensions[row_idx].height = top_margin_height
            ws.row_dimensions[row_idx+1].height = top_height
            ws.row_dimensions[row_idx+2].height = middle_margin_height
            ws.row_dimensions[row_idx+3].height = bottom_height
            ws.row_dimensions[row_idx+4].height = bottom_margin_height
        # インデクサの調整
        if col_idx == col_num:
            col_idx = 1
            row_idx += 5
        else:
            col_idx += 1
    # 幅の調整
    one_col_width = SHEET_A4_WIDTH // col_num
    for i in range(col_num):
        ws.column_dimensions[__convert_num2xlscol(2*i)].width = one_col_width * 3 / 10
        ws.column_dimensions[__convert_num2xlscol(2*i + 1)].width = one_col_width * 7 / 10

    return True


def __convert_num2xlscol(num):
    """概要(Summary line)

    intをエクセルの列文字に変換する
    0ならaとなるので、1列目はnumを0に設定

    Parameters
    ----------
    num : int
        列のインデクス番号

    Returns
    -------
    str
        エクセルの列文字

    """

    return chr(num + 97)


def __define_top_fontsize(num):
    """概要(Summary line)

    バーコードのフォントサイズを決める

    Parameters
    ----------
    num : int
        列数

    Returns
    -------
    int
        フォントサイズ

    """

    if num == 2:
        return 28   #26

    return 24   #11


def __define_bottom_fontsize(num):
    """概要(Summary line)

    バーコードデータ用文字のフォントサイズを決める

    Parameters
    ----------
    num : int
        列数

    Returns
    -------
    int
        フォントサイズ

    """

    if num == 2:
        return 14

    return 11