"""概要(Summary line)

初期位置・方位が不明な場合に推定を行う

Example
-------
python create_xlsx.py [f:filename]

"""

import os
import sys
import time

import numpy as np

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side


def create_xlsx_file(srcfile="out.csv", dstfile="out.xlsx"):
    """概要(Summary line)

    印刷用エクセルファイルを作成する

    Parameters
    ----------
    filename : str
        読み込み対象のcsvファイル名
    dstfile : str
        出力ファイル名

    Returns
    -------
    bool
        ファイル作成の成功or失敗

    """

    # ファイルの読み込み
    csv_data = []
    with open(srcfile, "r") as fr:
        line = fr.readline()
        while line:
            line_split = line.split(",")
            if len(line_split) != 4:
                return False
            if line_split[3][-1] == "\n":
                line_split[3] = line_split[3][:-1]
            csv_data.append([line_split[0], line_split[1], line_split[2], line_split[3]])
            line = fr.readline()
    # ワークブックと各種シートを作成
    wb = openpyxl.Workbook()
    if not __create_summary_sheet(wb, csv_data):
        return False
    if not __create_print_sheet(wb, csv_data):
        return False
    # 保存
    wb.active = wb.worksheets[0]
    if os.path.splitext(dstfile)[1] != ".xlsx":
        dstfile += ".xlsx"
    try:
        wb.save(dstfile)
    except(PermissionError):
        return False

    return True


def __create_summary_sheet(wb, csv_data):
    """概要(Summary line)

    要約リストのシートを作成する

    Parameters
    ----------
    wb : openpyxl.Workbook
        対象のワークブック
    csv_data : list
        csvファイルのデータ

    Returns
    -------
    bool
        シート作成の成功or失敗

    """

    # ヘッダー書き込み
    ws = wb.worksheets[0]
    ws.title = "バーコード全リスト"
    row_idx = 1 # 1オリジンのインデクサ
    cell_a = ws.cell(row_idx, 1)
    cell_b = ws.cell(row_idx, 2)
    cell_c = ws.cell(row_idx, 3)
    cell_d = ws.cell(row_idx, 4)
    cell_e = ws.cell(row_idx, 5)
    cell_a.value = "X"
    cell_b.value = "Y"
    cell_c.value = "Z"
    cell_d.value = "バーコードの値"
    cell_e.value = "バーコード"
    cell_a.alignment = Alignment(horizontal="right")
    cell_b.alignment = Alignment(horizontal="right")
    cell_c.alignment = Alignment(horizontal="right")
    cell_d.alignment = Alignment(horizontal="center")
    # バーコードリストデータ入力
    normal_alignment = Alignment(vertical="center")
    normal_font = Font(size=12)
    for csv_row in csv_data:
        row_idx += 1
        cell_a = ws.cell(row_idx, 1)
        cell_b = ws.cell(row_idx, 2)
        cell_c = ws.cell(row_idx, 3)
        cell_d = ws.cell(row_idx, 4)
        cell_e = ws.cell(row_idx, 5)
        cell_a.value = float(csv_row[0])
        cell_b.value = float(csv_row[1])
        cell_c.value = float(csv_row[2])
        cell_d.value = csv_row[3]
        cell_e.value = csv_row[3]
        cell_a.alignment = normal_alignment
        cell_b.alignment = normal_alignment
        cell_c.alignment = normal_alignment
        cell_d.alignment = Alignment(horizontal="center", vertical="center")
        cell_e.alignment = normal_alignment
        cell_a.font = normal_font
        cell_b.font = normal_font
        cell_c.font = normal_font
        cell_d.font = normal_font
        cell_e.font = Font(name="CODE39", size=14)
        ws.row_dimensions[row_idx].height = 30
    # セルの書式設定
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 31
    # ws.column_dimensions["D"].alignment = Alignment(horizontal="center")#できないっぽい
    # ws.column_dimensions["E"].font = Font(name="CODE39")#できないっぽい

    return True


def __create_print_sheet(wb, csv_data, col_num=2):
    """概要(Summary line)

    各バーコード専用のシートを作成する

    Parameters
    ----------
    wb : openpyxl.Workbook
        対象のワークブック
    csv_data : list
        csvファイルのデータ
    col_num : int
        バーコードの列数、最大26

    Returns
    -------
    bool
        シート作成の成功or失敗

    """

    # シートを作成
    wb.create_sheet("印刷用", 1)
    ws = wb.worksheets[1]
    # 列幅を設定
    WS_WIDTH = 80   # A4の場合
    col_width = WS_WIDTH // col_num
    for col_idx in range(col_num):
        ws.column_dimensions[__convert_num2xlscol(col_idx)].width = col_width
    # csvデータを入力
    row_idx = 1
    col_idx = 1
    alignment = Alignment(horizontal="center", vertical="center")
    top_font = Font(name="CODE39", size=__define_top_fontsize(col_num))
    bottom_font = Font(size=__define_bottom_fontsize(col_num))
    top_height = 36
    bottom_height = 20
    top_margin_height = 2
    middle_margin_height = 1
    bottom_margin_height = 1
    side = Side(style="thin")
    top_border = Border(top=side, left=side, right=side)
    middle_border = Border(left=side, right=side)
    bottom_border = Border(bottom=side, left=side, right=side)
    for csv_row in csv_data:
        # 値入力
        cell_top_margin = ws.cell(row_idx, col_idx)
        cell_top = ws.cell(row_idx+1, col_idx)
        cell_middle_margin = ws.cell(row_idx+2, col_idx)
        cell_bottom = ws.cell(row_idx+3, col_idx)
        cell_bottom_margin = ws.cell(row_idx+4, col_idx)
        cell_top.value = csv_row[3]
        cell_bottom.value = csv_row[3]
        # 書式設定
        cell_top.alignment = alignment
        cell_bottom.alignment = alignment
        cell_top.font = top_font
        cell_bottom.font = bottom_font
        cell_top_margin.border = top_border
        cell_top.border = middle_border
        cell_middle_margin.border = middle_border
        cell_bottom.border = middle_border
        cell_bottom_margin.border = bottom_border
        if col_idx == 1:
            ws.row_dimensions[row_idx].height = top_margin_height
            ws.row_dimensions[row_idx+1].height = top_height
            ws.row_dimensions[row_idx+2].height = middle_margin_height
            ws.row_dimensions[row_idx+3].height = bottom_height
            ws.row_dimensions[row_idx+4].height = bottom_margin_height
        # インデクサの調整
        if col_idx == col_num:
            col_idx = 1
            row_idx += 5
        else:
            col_idx += 1

    return True


def __convert_num2xlscol(num):
    """概要(Summary line)

    intをエクセルの列文字に変換する

    Parameters
    ----------
    num : int
        列のインデクス番号

    Returns
    -------
    str
        エクセルの列文字

    """

    return chr(num + 97)


def __define_top_fontsize(num):
    """概要(Summary line)

    バーコードのフォントサイズを決める

    Parameters
    ----------
    num : int
        列数

    Returns
    -------
    int
        フォントサイズ

    """

    if num == 2:
        return 26

    return 11


def __define_bottom_fontsize(num):
    """概要(Summary line)

    バーコードデータ用文字のフォントサイズを決める

    Parameters
    ----------
    num : int
        列数

    Returns
    -------
    int
        フォントサイズ

    """

    if num == 2:
        return 14

    return 11


# メイン文の実行
if __name__ == "__main__":
    if len(sys.argv) < 2:
        result = create_xlsx_file()
    else:
        result = create_xlsx_file(sys.argv[1])
    if not result:
        print("失敗")
